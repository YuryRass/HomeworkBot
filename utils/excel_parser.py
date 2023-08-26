"""
    Модуль excel_parser.py осуществляет парсинг excel файла
    и извлекает соответвующие данные о студентах и преподах.
"""
from enum import Enum

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from model.pydantic.db_start_data import StudentRaw, TeacherRaw


class ExcelDataParserError(Exception):
    """Возможные исключения при парсинге excel файла."""
    pass


class ParserType(Enum):
    """
        Класс возможных способов парсинга excel файла.
        Атрибуты:
        ALL - парсинг данных про студентов и преподов.
        TEACHER - парсинг данных про преподов.
        STUDENT - парсинг данных про студнтов.
    """
    ALL = 0
    TEACHER = 1
    STUDENT = 2


class ExcelDataParser:
    """Парсер данных их excel файла"""

    def __init__(self, file_path: str, parse_type: ParserType = ParserType.ALL):
        """
            Параметры:
            file_path (str): путь до excel файла.
            parse_type (ParserType): тип парсинга файла.
        """
        # В качестве ключей в словарях self.__student и self.__teacher
        # используются названия учебных дисциплин. Ключами для словарей
        # dict[str, list[StudentRaw]] и dict[str, list[TeacherRaw]]
        # используются названия учебных групп.
        self.__student: dict[str, dict[str, list[StudentRaw]]] = {}
        self.__teacher: dict[str, dict[str, list[TeacherRaw]]] = {}

        self.__parse_type = parse_type
        # инициализация переменных __student и __teacher
        # из excel файла
        self.__load_data(file_path, parse_type)

    @property
    def students(self) -> dict[str, dict[str, list[StudentRaw]]]:
        """Возвращает словарь с данными о студентах"""
        if self.__parse_type == ParserType.TEACHER:
            raise ExcelDataParserError(
                "Students data don't with this ParseType"
            )
        return self.__student

    @property
    def teachers(self) -> dict[str, dict[str, list[TeacherRaw]]]:
        """Возвращает словарь с данными о преподах"""
        if self.__parse_type == ParserType.STUDENT:
            raise ExcelDataParserError(
                "Teachers data don't with this ParseType"
            )
        return self.__teacher

    def __load_data(self, file_path: str, parse_type: ParserType) -> None:
        """
            Загрузка данных из excel файла
            Параметры:
            file_path (str): путь до excel файла.
            parse_type (ParserType): способ парсинга.
        """
        # получаем excel книгу с рабочими листами
        wb: Workbook = openpyxl.load_workbook(file_path)

        # Осуществляем парсинг excel листов согласно
        # значению параметра parse_type.
        # 'index = wb.sheetnames.index(<name_work_sheet>)' -
        # получение номера индекса для указанного excel листа.
        # 'wb.active = index' - активация excel листа,
        # индекс которого равен значению переменой index.
        match parse_type:
            case ParserType.ALL:
                index = wb.sheetnames.index('teachers')
                wb.active = index
                self.__teachers_parser(wb.active)
                index = wb.sheetnames.index('students')
                wb.active = index
                self.__students_parser(wb.active)
            case ParserType.STUDENT:
                index = wb.sheetnames.index('students')
                wb.active = index
                self.__students_parser(wb.active)
            case ParserType.TEACHER:
                index = wb.sheetnames.index('teachers')
                wb.active = index
                self.__teachers_parser(wb.active)
            case _:
                raise ExcelDataParserError('ParserType not found')

    def __teachers_parser(self, worksheet: Worksheet) -> None:
        """
            Парсинг excel страницы про преподавателей.
            Параметры:
            worksheet (Worksheet): excel страница.
        """
        teacher_name = ''
        # начинаем со второй строки, исключая первую строку
        # с заголовками.
        row = 2

        # построчно читаем excel страницу о преподавателях,
        # пока не дойдем до пустой строки, где teacher_name = None
        while teacher_name is not None:
            # возвращаем значения excel ячеек по указанным координатам
            teacher_name = worksheet.cell(row=row, column=1).value
            telegram_id = worksheet.cell(row=row, column=2).value
            discipline = worksheet.cell(row=row, column=3).value
            is_admin = bool(worksheet.cell(row=row, column=4).value)
            group = worksheet.cell(row=row, column=5).value

            # условие закрытия цикла
            if teacher_name is None:
                break

            # добавление новой учебной дисциплины
            if discipline not in self.__teacher:
                self.__teacher[discipline] = {}

            # добавление новой учебной группы
            if group not in self.__teacher[discipline]:
                self.__teacher[discipline][group] = []

            # добавление данных о преподах
            # (ФИО, Telegram ID, флаг на админа (True or False))
            self.__teacher[discipline][group].append(
                TeacherRaw(
                    full_name=teacher_name,
                    telegram_id=telegram_id,
                    is_admin=is_admin
                )
            )

            # переход на след. строку
            row += 1

    def __students_parser(self, worksheet: Worksheet) -> None:
        """
            Парсинг excel страницы про студентов.
            Параметры:
            worksheet (Worksheet): excel страница.
        """
        student_name = ''
        # начинаем со второй строки, исключая первую строку
        # с заголовками.
        row = 2

        # построчно читаем excel страницу о студентах,
        # пока не дойдем до пустой строки, где student_name = None
        while student_name is not None:
            # возвращаем значения excel ячеек по указанным координатам
            student_name = worksheet.cell(row=row, column=1).value
            group = worksheet.cell(row=row, column=2).value
            discipline = worksheet.cell(row=row, column=3).value

            # условие закрытия цикла
            if student_name is None:
                break

            # добавление новой дисциплины
            if discipline not in self.__student:
                self.__student[discipline] = {}

            # добавление новой учебной группы
            if group not in self.__student[discipline]:
                self.__student[discipline][group] = []

            # добавление ФИО студентов
            self.__student[discipline][group].append(
                StudentRaw(
                    full_name=student_name,
                )
            )

            # переход на след. строку
            row += 1
