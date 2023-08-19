"""
Модуль реализует базовый отчет по результатам выполненных домашних работ у студентов.
На основе данного отчета формируются другие отчеты.
"""
import json
from datetime import datetime
from enum import IntEnum
from pathlib import Path

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from database.main_db import common_crud
from model.pydantic.home_work import DisciplineHomeWorks

from config import settings


class ReportFieldEnum(IntEnum):
    """
    Номер столбца отчета в формируемом файле
    """
    STUDENT_NAME = 1
    POINTS = 2
    LAB_COMPLETED = 3
    DEADLINES_FAILS = 4
    TASKS_COMPLETED = 5
    TASK_RATIO = 6
    NEXT = 7


class BaseReportBuilder:
    """
    Базовый класс, закладывающий каркас отчета

    Атрибуты:

    group_id: ID учбеной группы.

    discipline_id: ID дисицплины

    group_name: имя группы.

    discipline_name: название дисциплины.

    tasks_in_discipline: число заданий в дисциплине.

    __file_path: полное название файла отчета.

    wb: excel книга.
    """

    # базовые цвета для отчета.
    RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
    GREEN_FILL = PatternFill(start_color='006633', end_color='006633', fill_type="solid")

    def __init__(self, group_id: int, discipline_id: int,
                 prefix_file: str, extension: str = 'xlsx'):
        """
        :param group_id: идентификатор группы
        :param discipline_id: идентификатор дисциплины
        :param prefix_file: префикс имени формируемого отчета
        :param extension: расширение файла отчета
        :return:
        """
        self.group_id = group_id
        self.discipline_id = discipline_id

        group = common_crud.get_group(group_id)
        discipline = common_crud.get_discipline(discipline_id)

        self.group_name = group.group_name
        self.discipline_name = discipline.short_name
        self.tasks_in_discipline = discipline.max_tasks

        path = Path(Path.cwd().joinpath(settings.TEMP_REPORT_DIR))
        self.__file_path = Path(
            path.joinpath(f'{discipline.short_name}_{prefix_file}_{group.group_name}.{extension}')
        )

        self.wb = Workbook()

    def build_report(self) -> None:
        """
        Метод запускающий создание и заполнение базового отчета

        :return: None
        """

        # инициализируем рабочую страницу
        worksheet = self.wb.active
        worksheet.title = self.discipline_name

        students = common_crud.get_students_from_group(self.group_id)
        row = 1
        for student in students:
            # результаты по выполнению домашних работ
            answers = common_crud.get_student_discipline_answer(student.id, self.discipline_id)
            home_works = DisciplineHomeWorks(**json.loads(answers.home_work)).home_works

            # делаем заголовок в excel странице
            if row == 1:
                worksheet.cell(row=row, column=ReportFieldEnum.STUDENT_NAME).value = 'ФИО студента'
                worksheet.cell(row=row, column=ReportFieldEnum.POINTS).value = 'Баллы (макс. 100)'
                worksheet.cell(row=row, column=ReportFieldEnum.LAB_COMPLETED).value = 'Полностью выполненых лаб'
                worksheet.cell(row=row, column=ReportFieldEnum.DEADLINES_FAILS).value = 'Кол-во сорванных дедлайнов'

                worksheet.cell(row=row, column=ReportFieldEnum.TASKS_COMPLETED).value = 'Кол-во выполненых задач'
                worksheet.cell(row=row, column=ReportFieldEnum.TASK_RATIO).value = 'task ratio'
                row += 1
            if row > 1:
                worksheet.cell(row=row, column=ReportFieldEnum.STUDENT_NAME).value = student.full_name
                worksheet.cell(row=row, column=ReportFieldEnum.POINTS).value = answers.point

                deadlines_fails = 0 # кол-во просроченных дедлайнов
                lab_completed = 0 # кол-во выполненных лаб. работ
                for work in home_works:
                    if work.is_done:
                        lab_completed += 1
                    if datetime.now().date() > work.deadline:
                        if work.end_time is not None:
                            if work.end_time.date() > work.deadline:
                                deadlines_fails += 1
                        else:
                            deadlines_fails += 1
                worksheet.cell(row=row, column=ReportFieldEnum.LAB_COMPLETED).value = lab_completed
                worksheet.cell(row=row, column=ReportFieldEnum.DEADLINES_FAILS).value = deadlines_fails

                task_completed = 0 # кол-во выполненных задач

                for number_lab, work in enumerate(home_works):
                    for number_task, task in enumerate(work.tasks):
                        if task.is_done:
                            task_completed += 1

                worksheet.cell(row=row, column=ReportFieldEnum.TASKS_COMPLETED).value = task_completed

                worksheet.cell(
                    row=row, column=ReportFieldEnum.TASK_RATIO
                ).value = task_completed / self.tasks_in_discipline

            row += 1

    def save_report(self):
        """
        Метод сохранения отчета

        :return: None
        """
        self.wb.save(self.get_path_to_report())

    def get_path_to_report(self) -> str:
        """
        :return: путь до сформированного отчета
        """
        return str(self.__file_path)